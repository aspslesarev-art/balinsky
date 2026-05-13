#!/usr/bin/env python3
"""
estatemarket.io occupancy + ADR aggregator for every Bali complex.

For each row in raw_complexes (with parseable Geo / Geo 2), pulls every
estatemarket.io listing within 500 m, fetches its card to learn the
type + bedroom count, and aggregates villa-vs-apartment metrics.

Outputs an Excel report (two sheets, "Сводка" + "Детализация") and,
unless `--no-supabase`, upserts the per-complex aggregates into a new
`complex_market_stats` table for catalog use later.

Usage:
  python scripts/estatemarket_occupancy.py
  python scripts/estatemarket_occupancy.py --input complexes.csv
  python scripts/estatemarket_occupancy.py --no-supabase --output report.xlsx
"""

import argparse
import asyncio
import json
import math
import os
import re
import sys
import time
from pathlib import Path
from typing import Any

import requests

INDEX_URL = "https://estatemarket.io/api/booking_data-map"
CARD_URL  = "https://estatemarket.io/api/booking_data_map_card/{id}"

# Persistent on-disk cache so repeat runs are cheap. Cards rarely change
# day-to-day, occupancy refreshes maybe weekly. Bump the schema number
# in the file if format changes.
CACHE_PATH = Path(__file__).parent / "estatemarket_cards.json"

RADIUS_M = 1000
EARTH_R  = 6_371_000

# Types worth counting. estatemarket sometimes uses "Villa" (singular)
# and "Villas" (plural) interchangeably, so we treat both as villas.
VILLA_TYPES = {"villa", "villas"}
APT_TYPES   = {"apartments", "apartment"}


def haversine_m(lat1: float, lng1: float, lat2: float, lng2: float) -> float:
    dLa = math.radians(lat2 - lat1)
    dLo = math.radians(lng2 - lng1)
    a = (math.sin(dLa / 2) ** 2
         + math.cos(math.radians(lat1)) * math.cos(math.radians(lat2))
         * math.sin(dLo / 2) ** 2)
    return 2 * EARTH_R * math.asin(math.sqrt(a))


def load_env_local() -> None:
    p = Path(".env.local")
    if not p.exists():
        return
    for line in p.read_text().splitlines():
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        if k.strip() not in os.environ:
            os.environ[k.strip()] = v.strip().strip('"').strip("'")


def parse_geo(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, list):
        for item in v:
            r = parse_geo(item)
            if r is not None:
                return r
        return None
    s = str(v).replace(",", ".").strip()
    m = re.search(r"-?\d+(?:\.\d+)?", s)
    if not m:
        return None
    try:
        f = float(m.group(0))
        return f if f != 0 else None
    except ValueError:
        return None


KIND_TO_RAW = {
    "complex":   "raw_complexes",
    "villa":     "raw_villas",
    "apartment": "raw_apartments",
}
KIND_TO_DEST = {
    "complex":   "complex_market_stats",
    "villa":     "villa_market_stats",
    "apartment": "apartment_market_stats",
}


def load_listings_from_supabase(kind: str) -> list[dict]:
    """Pull every published row of the given kind from Supabase and
    return the minimal shape estatemarket_occupancy needs.
    Complexes are gated by Статус, villas/apartments by `Опубликовать`."""
    SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
    SB_KEY = os.environ["SUPABASE_SERVICE_KEY"]
    h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}"}
    raw_table = KIND_TO_RAW[kind]
    out: list[dict] = []
    PAGE = 200
    for offset in range(0, 8000, PAGE):
        r = requests.get(
            f"{SB_URL}/rest/v1/{raw_table}?select=airtable_id,data&limit={PAGE}&offset={offset}",
            headers=h, timeout=30,
        )
        r.raise_for_status()
        batch = r.json()
        if not batch:
            break
        for row in batch:
            d = row.get("data") or {}
            if kind == "complex":
                status = str(d.get("Статус") or "").lower()
                if not status or "архив" in status or "не публик" in status:
                    continue
            else:
                if d.get("Опубликовать") is not True:
                    continue
            lat = parse_geo(d.get("Geo"))
            lon = parse_geo(d.get("Geo 2"))
            if lat is None or lon is None:
                continue
            seo_title = d.get("SEO:Title")
            if isinstance(seo_title, dict):
                seo_title = seo_title.get("value")
            name = d.get("Project") or seo_title or d.get("Name") or d.get("Название") or row["airtable_id"]
            out.append({
                "airtable_id": row["airtable_id"],
                "name": str(name),
                "lat": lat, "lng": lon,
                "types": d.get("Типы юнитов"),
                "developer": d.get("Developer1") or d.get("Варианты поиска застройщика"),
            })
        if len(batch) < PAGE:
            break
    return out


def load_complexes_from_supabase() -> list[dict]:
    return load_listings_from_supabase("complex")


def load_complexes_from_csv(path: str) -> list[dict]:
    import pandas as pd
    df = pd.read_csv(path) if path.lower().endswith(".csv") else pd.read_excel(path)
    out: list[dict] = []
    for _, row in df.iterrows():
        lat = parse_geo(row.get("lat") if "lat" in df.columns else None)
        lng = parse_geo(row.get("lng") if "lng" in df.columns else (row.get("lon") if "lon" in df.columns else None))
        if lat is None or lng is None:
            continue
        out.append({
            "airtable_id": str(row.get("airtable_id") or row.get("name") or len(out)),
            "name": str(row.get("name") or "unknown"),
            "lat": lat, "lng": lng,
            "types": row.get("type"),
            "developer": None,
        })
    return out


def fetch_index() -> list[dict]:
    print("fetching estatemarket index …", file=sys.stderr)
    r = requests.post(
        INDEX_URL,
        json={"status": None, "country": None, "city": None, "type": None, "facilities": None, "price": None},
        headers={
            "Content-Type": "application/json",
            "Accept": "application/json",
            "X-Requested-With": "XMLHttpRequest",
        },
        timeout=60,
    )
    r.raise_for_status()
    data = r.json()
    print(f"got {len(data)} listings", file=sys.stderr)
    # Parse coords once
    out: list[dict] = []
    for d in data:
        loc = d.get("location") or []
        if not isinstance(loc, list) or len(loc) != 2:
            continue
        try:
            lat = float(loc[0]); lng = float(loc[1])
        except (TypeError, ValueError):
            continue
        out.append({
            "id": d["id"],
            "title": d.get("title"),
            "occupancy": d.get("occupancy"),
            "price": d.get("price"),
            "lat": lat,
            "lng": lng,
        })
    return out


def load_cache() -> dict[str, dict]:
    if not CACHE_PATH.exists():
        return {}
    try:
        data = json.loads(CACHE_PATH.read_text())
        if not isinstance(data, dict):
            return {}
        return data
    except json.JSONDecodeError:
        return {}


def save_cache(cache: dict[str, dict]) -> None:
    tmp = CACHE_PATH.with_suffix(".tmp")
    tmp.write_text(json.dumps(cache, ensure_ascii=False))
    tmp.replace(CACHE_PATH)


async def fetch_cards(ids: list[int], cache: dict[str, dict], concurrency: int = 8) -> None:
    """Fill the cache for any id not already present. Saves every 200 fetches."""
    import aiohttp
    missing = [i for i in ids if str(i) not in cache]
    if not missing:
        return
    print(f"fetching {len(missing)} cards (cache has {len(cache)}) …", file=sys.stderr)
    sem = asyncio.Semaphore(concurrency)
    fetched_since_save = 0

    async def one(session: "aiohttp.ClientSession", lid: int) -> None:
        nonlocal fetched_since_save
        async with sem:
            try:
                async with session.get(CARD_URL.format(id=lid), timeout=aiohttp.ClientTimeout(total=30)) as r:
                    if r.status != 200:
                        cache[str(lid)] = {"_error": r.status}
                        return
                    j = await r.json()
                    cache[str(lid)] = {
                        "type": j.get("type"),
                        "rooms": len(j.get("rooms") or []),
                        "occupancy": j.get("occupancy"),
                        "min_price": j.get("min_price"),
                        "max_price": j.get("max_price"),
                    }
            except Exception as e:
                cache[str(lid)] = {"_error": str(e)[:80]}

    async with aiohttp.ClientSession(
        headers={"Accept": "application/json", "User-Agent": "Mozilla/5.0 (balinsky-occupancy)"},
    ) as session:
        # Batch in chunks so we can sleep between batches + save cache progressively.
        BATCH = 40
        for start in range(0, len(missing), BATCH):
            chunk = missing[start: start + BATCH]
            await asyncio.gather(*(one(session, lid) for lid in chunk))
            fetched_since_save += len(chunk)
            if fetched_since_save >= 200:
                save_cache(cache)
                fetched_since_save = 0
                print(f"  cached {len(cache)} so far", file=sys.stderr)
            await asyncio.sleep(0.2)
    save_cache(cache)


def aggregate(listings_in_radius: list[dict], cache: dict[str, dict]) -> dict:
    """Per-complex summary: counts, occupancy avg, ADR avg, RevPAR. Each
    bucket only counts entries where the metric is non-null."""
    villa_occ: list[float] = []
    villa_price: list[float] = []
    villa_count = 0
    apt_occ: list[float] = []
    apt_price: list[float] = []
    apt_count = 0

    def f_or_none(v: Any) -> float | None:
        if v is None:
            return None
        try:
            return float(v)
        except (TypeError, ValueError):
            return None

    for ls in listings_in_radius:
        card = cache.get(str(ls["id"]))
        if not card or "_error" in card:
            continue
        kind = (card.get("type") or "").lower().strip()
        occ = f_or_none(ls.get("occupancy") or card.get("occupancy"))
        price = f_or_none(ls.get("price"))
        # If index price is null but card has min/max, use the average.
        if price is None:
            lo, hi = f_or_none(card.get("min_price")), f_or_none(card.get("max_price"))
            if lo is not None and hi is not None:
                price = (lo + hi) / 2
            elif lo is not None:
                price = lo
            elif hi is not None:
                price = hi
        if kind in VILLA_TYPES:
            villa_count += 1
            if occ is not None: villa_occ.append(occ)
            if price is not None: villa_price.append(price)
        elif kind in APT_TYPES:
            apt_count += 1
            if occ is not None: apt_occ.append(occ)
            if price is not None: apt_price.append(price)

    def mean(arr):
        return round(sum(arr) / len(arr), 2) if arr else None
    def stable_mean(arr):
        return mean(arr) if len(arr) >= 3 else None

    villa_occ_avg = mean(villa_occ)
    apt_occ_avg = mean(apt_occ)
    villa_adr = stable_mean(villa_price)
    apt_adr = stable_mean(apt_price)

    return {
        "total_listings": len(listings_in_radius),
        "villa_count": villa_count,
        "villa_occ_avg": villa_occ_avg,
        "villa_adr": villa_adr,
        "villa_revpar": (villa_occ_avg / 100 * villa_adr) if (villa_occ_avg is not None and villa_adr is not None) else None,
        "apt_count": apt_count,
        "apt_occ_avg": apt_occ_avg,
        "apt_adr": apt_adr,
        "apt_revpar": (apt_occ_avg / 100 * apt_adr) if (apt_occ_avg is not None and apt_adr is not None) else None,
    }


def write_excel(out_path: str, complexes: list[dict], summary: dict[str, dict], details: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    head_font = Font(name="Arial", size=11, bold=True, color="FFFFFF")
    body_font = Font(name="Arial", size=11)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    center = Alignment(horizontal="left", vertical="center")

    # --- Sheet 1: Сводка ---
    ws1 = wb.active
    ws1.title = "Сводка"
    cols = [
        ("ЖК", 32), ("Lat", 11), ("Lng", 11), ("Типы", 18), ("Застройщик", 22),
        ("Виллы, шт", 11), ("Виллы, occ %", 13), ("Виллы, ADR $", 13), ("Виллы, RevPAR $", 15),
        ("Апарт., шт", 11), ("Апарт., occ %", 13), ("Апарт., ADR $", 13), ("Апарт., RevPAR $", 15),
        ("Листингов в 500м", 16),
    ]
    for i, (label, w) in enumerate(cols, 1):
        cell = ws1.cell(row=1, column=i, value=label)
        cell.font = head_font; cell.fill = header_fill; cell.alignment = center
        ws1.column_dimensions[get_column_letter(i)].width = w

    for r, cx in enumerate(complexes, 2):
        s = summary.get(cx["airtable_id"], {})
        row = [
            cx["name"], cx["lat"], cx["lng"],
            cx.get("types") if not isinstance(cx.get("types"), list) else ", ".join(map(str, cx["types"])),
            cx.get("developer") if not isinstance(cx.get("developer"), list) else ", ".join(map(str, cx["developer"])),
            s.get("villa_count"), s.get("villa_occ_avg"), s.get("villa_adr"), None,   # RevPAR via formula
            s.get("apt_count"),   s.get("apt_occ_avg"),   s.get("apt_adr"),   None,
            s.get("total_listings"),
        ]
        for c, val in enumerate(row, 1):
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = body_font
            cell.alignment = center
        # RevPAR formulas (columns 7=occ%, 8=ADR, 9=RevPAR)  villa, then 11/12/13 for apt
        if s.get("villa_occ_avg") is not None and s.get("villa_adr") is not None:
            ws1.cell(row=r, column=9, value=f"=G{r}/100*H{r}").number_format = '"$"#,##0'
        if s.get("apt_occ_avg") is not None and s.get("apt_adr") is not None:
            ws1.cell(row=r, column=13, value=f"=K{r}/100*L{r}").number_format = '"$"#,##0'
        # Percent + currency formats
        for c in (7, 11):
            ws1.cell(row=r, column=c).number_format = '0.0"%"'
        for c in (8, 12):
            ws1.cell(row=r, column=c).number_format = '"$"#,##0'

    ws1.freeze_panes = "A2"

    # --- Sheet 2: Детализация ---
    ws2 = wb.create_sheet("Детализация")
    det_cols = [
        ("ЖК", 30), ("ID", 8), ("Тип", 14), ("Спальни", 9),
        ("Occupancy %", 13), ("Цена/ночь $", 13), ("Дист., м", 10), ("Ссылка", 50),
    ]
    for i, (label, w) in enumerate(det_cols, 1):
        cell = ws2.cell(row=1, column=i, value=label)
        cell.font = head_font; cell.fill = header_fill; cell.alignment = center
        ws2.column_dimensions[get_column_letter(i)].width = w
    for r, d in enumerate(details, 2):
        ws2.cell(row=r, column=1, value=d["complex_name"]).font = body_font
        ws2.cell(row=r, column=2, value=d["id"]).font = body_font
        ws2.cell(row=r, column=3, value=d["type"]).font = body_font
        ws2.cell(row=r, column=4, value=d["rooms"]).font = body_font
        c5 = ws2.cell(row=r, column=5, value=d["occupancy"])
        c5.font = body_font; c5.number_format = '0.0"%"'
        c6 = ws2.cell(row=r, column=6, value=d["price"])
        c6.font = body_font; c6.number_format = '"$"#,##0'
        ws2.cell(row=r, column=7, value=round(d["distance_m"])).font = body_font
        url = f"https://estatemarket.io/booking_data-map?id={d['id']}"
        ws2.cell(row=r, column=8, value=url).font = body_font
    ws2.freeze_panes = "A2"

    wb.save(out_path)


def upsert_supabase(listings: list[dict], summary: dict[str, dict], kind: str = "complex") -> None:
    """Persist per-listing aggregates to <kind>_market_stats."""
    SB_URL = os.environ["NEXT_PUBLIC_SUPABASE_URL"]
    SB_KEY = os.environ["SUPABASE_SERVICE_KEY"]
    h = {"apikey": SB_KEY, "Authorization": f"Bearer {SB_KEY}",
         "Content-Type": "application/json",
         "Prefer": "resolution=merge-duplicates,return=minimal"}

    dest_table = KIND_TO_DEST[kind]
    rows = []
    for cx in listings:
        s = summary.get(cx["airtable_id"]) or {}
        rows.append({
            "airtable_id": cx["airtable_id"],
            "lat": cx["lat"], "lon": cx["lng"],
            "total_listings_500m": s.get("total_listings", 0),
            "villa_count": s.get("villa_count", 0),
            "villa_occupancy_pct": s.get("villa_occ_avg"),
            "villa_adr_usd": s.get("villa_adr"),
            "villa_revpar_usd": s.get("villa_revpar"),
            "apartment_count": s.get("apt_count", 0),
            "apartment_occupancy_pct": s.get("apt_occ_avg"),
            "apartment_adr_usd": s.get("apt_adr"),
            "apartment_revpar_usd": s.get("apt_revpar"),
        })

    BATCH = 200
    for off in range(0, len(rows), BATCH):
        chunk = rows[off: off + BATCH]
        r = requests.post(
            f"{SB_URL}/rest/v1/{dest_table}?on_conflict=airtable_id",
            json=chunk, headers=h, timeout=30,
        )
        if r.status_code not in (200, 201, 204):
            print(f"  upsert failed: {r.status_code} {r.text[:200]}", file=sys.stderr)
            return
    print(f"  upserted {len(rows)} rows → {dest_table}", file=sys.stderr)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--kind", choices=["villa", "apartment", "complex", "all"], default="complex",
                    help="which raw_* table to pull listings from")
    ap.add_argument("--input", help="CSV/XLSX with name,lat,lng (default: pull from Supabase)")
    ap.add_argument("--output", default="estatemarket_occupancy_report.xlsx")
    ap.add_argument("--limit", type=int, default=0)
    ap.add_argument("--no-supabase", action="store_true", help="skip writing market_stats tables")
    args = ap.parse_args()

    load_env_local()

    if args.kind == "all":
        for k in ("complex", "villa", "apartment"):
            args.kind = k
            run_one(args)
        return
    run_one(args)


def run_one(args) -> None:
    if args.input:
        listings = load_complexes_from_csv(args.input)
    else:
        listings = load_listings_from_supabase(args.kind)
    if args.limit:
        listings = listings[: args.limit]
    print(f"loaded {len(listings)} {args.kind}s", file=sys.stderr)
    complexes = listings  # alias to keep the rest of the body unchanged

    index = fetch_index()

    # First pass: for each complex collect listings within 500 m. Build
    # the union of all needed listing ids — we only fetch cards for those.
    per_complex_listings: dict[str, list[dict]] = {}
    needed_ids: set[int] = set()
    for cx in complexes:
        clat, clng = cx["lat"], cx["lng"]
        # Cheap pre-filter: bounding-box first, haversine only on survivors.
        # 0.011° ≈ 1.2 km at the equator — easily covers 1000m.
        BOX = 0.011
        candidates = [
            ls for ls in index
            if abs(ls["lat"] - clat) < BOX and abs(ls["lng"] - clng) < BOX
        ]
        in_radius: list[dict] = []
        for ls in candidates:
            d = haversine_m(clat, clng, ls["lat"], ls["lng"])
            if d <= RADIUS_M:
                ls2 = dict(ls); ls2["distance_m"] = d
                in_radius.append(ls2)
                needed_ids.add(ls["id"])
        per_complex_listings[cx["airtable_id"]] = in_radius
    print(f"need cards for {len(needed_ids)} listings", file=sys.stderr)

    # Second pass: fetch cards for everything we need, with cache.
    cache = load_cache()
    asyncio.run(fetch_cards(sorted(needed_ids), cache))

    # Third pass: aggregate.
    summary: dict[str, dict] = {}
    details: list[dict] = []
    for cx in complexes:
        listings = per_complex_listings[cx["airtable_id"]]
        s = aggregate(listings, cache)
        summary[cx["airtable_id"]] = s
        for ls in listings:
            card = cache.get(str(ls["id"]), {})
            kind = (card.get("type") or "").lower().strip()
            if kind not in VILLA_TYPES and kind not in APT_TYPES:
                continue
            price = ls.get("price")
            if price is None:
                lo, hi = card.get("min_price"), card.get("max_price")
                if lo is not None and hi is not None:
                    price = (lo + hi) / 2
                elif lo is not None:
                    price = lo
                elif hi is not None:
                    price = hi
            details.append({
                "complex_name": cx["name"],
                "id": ls["id"],
                "type": card.get("type"),
                "rooms": card.get("rooms"),
                "occupancy": float(ls["occupancy"]) if ls.get("occupancy") not in (None, "") else None,
                "price": price,
                "distance_m": ls["distance_m"],
            })

    out_path = args.output
    if args.kind != "complex" and out_path == "estatemarket_occupancy_report.xlsx":
        out_path = f"estatemarket_occupancy_{args.kind}.xlsx"
    write_excel(out_path, complexes, summary, details)
    print(f"wrote {out_path}", file=sys.stderr)

    if not args.no_supabase:
        try:
            upsert_supabase(complexes, summary, args.kind)
        except KeyError as e:
            print(f"  Supabase env missing ({e}) — skipping upsert", file=sys.stderr)

    # QC summary
    total_villas = sum(s.get("villa_count", 0) for s in summary.values())
    total_apts = sum(s.get("apt_count", 0) for s in summary.values())
    empty = sum(1 for s in summary.values() if s.get("villa_count", 0) == 0 and s.get("apt_count", 0) == 0)
    print("", file=sys.stderr)
    print(f"  {args.kind}s processed:    {len(complexes)}", file=sys.stderr)
    print(f"  villas found total:       {total_villas}", file=sys.stderr)
    print(f"  apartments found total:   {total_apts}", file=sys.stderr)
    print(f"  {args.kind}s with 0 hits:  {empty}", file=sys.stderr)
    print(f"  report:                   {out_path}", file=sys.stderr)


if __name__ == "__main__":
    main()
